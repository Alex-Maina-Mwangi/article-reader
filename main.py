'''
 The script below utilizes europepmc Articles RESTful API
 https://europepmc.org/RestfulWebService to scrape through the fulltext of
 a subset of peer-reviewed articles with the sole objective of knowing the country
 and in the case of Kenya, the county where the study sites were located.

 Alex Maina - Librarian KEMRI-Wellcome Trust Research Programme

'''

import requests
from bs4 import BeautifulSoup
import re
import geonamescache
import spacy
from spacy.matcher import PhraseMatcher
import pycountry
from openpyxl import Workbook



pmcids = []



data = []
for i in pmcids:
    url = 'https://www.ebi.ac.uk/europepmc/webservices/rest/'+i+'/fullTextXML'

    xml_data = requests.get(url).content
    soup = BeautifulSoup(xml_data, 'xml')
    #get PMID
    pmid_tag = soup.find("article-id", {"pub-id-type": "pmid"})

    # Extract the PMID if not exist return None
    pmid = pmid_tag.text if pmid_tag else None


    sections = []
    for sec in soup.find_all("sec"):
        sec_data = {
            "id": sec.get("id"),
            "sec_type": sec.get("sec-type"),
            "title": sec.find("title").text if sec.find("title") else None,
            "content": sec.find("p").text.strip() if sec.find("p") else None
        }
        sections.append(sec_data)

    # Print structured data
    for section in sections:
        if section['title'] == 'Methods':
            paragraph = section['content']
            country_names = re.findall(r'\((.*?)\)', paragraph)
            gc = geonamescache.GeonamesCache()
            countries = gc.get_countries()
            country_names = [country['name'] for country in countries.values()]

            # Extract mentioned country names
            mentioned_countries = [country for country in country_names if country in paragraph]

            kenyan_counties = [ "Baringo", "Bomet", "Bungoma", "Busia", "Elgeyo-Marakwet", "Embu", 
                        "Garissa", "Homa Bay", "Isiolo", "Kajiado", "Kakamega", "Kericho", 
            "Kiambu", "Kilifi", "Kirinyaga", "Kisii", "Kisumu", "Kitui", "Kwale", 
            "Laikipia", "Lamu", "Machakos", "Makueni", "Mandera", "Marsabit", 
            "Meru", "Migori", "Mombasa", "Murang'a", "Nairobi", "Nakuru", 
            "Nandi", "Narok", "Nyamira", "Nyandarua", "Nyeri", "Samburu", 
            "Siaya", "Taita-Taveta", "Tana River", "Tharaka-Nithi", "Trans Nzoia", 
            "Turkana", "Uasin Gishu", "Vihiga", "Wajir", "West Pokot"
                ]

            # Load spaCy's English language model
            nlp = spacy.load("en_core_web_sm")

            # Create a PhraseMatcher and add counties
            matcher = PhraseMatcher(nlp.vocab)
            patterns = [nlp.make_doc(county) for county in kenyan_counties]
            matcher.add("KenyanCounties", patterns)

            def find_counties_spacy(text, matcher):
                doc = nlp(text)
                matches = matcher(doc)
                found_counties = set([doc[start:end].text for match_id, start, end in matches])
                found_counties = list(found_counties)
                return found_counties

            #find the counties
            found_counties = find_counties_spacy(paragraph, matcher)
            ##print("Found Counties:", found_counties)
            output = [[pmid],mentioned_countries,found_counties,[url]]
    #Append the output as a list named data
    data.append(output)

#print(data)
# Create a new Excel workbook
workbook = Workbook()
sheet = workbook.active
sheet.title = "List Data"

# Add column headers
sheet.cell(row=1, column=1, value="PMID")
sheet.cell(row=1, column=2, value="Countries")
sheet.cell(row=1, column=3, value="Counties")
sheet.cell(row=1, column=4, value="url")

# Write the data to Excel starting from the second row
row = 2
for record in data:
    col = 1
    for sublist in record:
        cleaned_sublist = [str(item) if item is not None else "" for item in sublist]
 
        sheet.cell(row=row, column=col, value=", ".join(cleaned_sublist))
        col += 1
    row += 1

# Save the workbook
output_file = "new_list_data_with_headers.xlsx"
workbook.save(output_file)

print(f"Data with headers saved to {output_file}!")

    
